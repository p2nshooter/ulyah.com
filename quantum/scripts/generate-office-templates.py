#!/usr/bin/env python3
"""
Membuat template Excel berkop surat CV. Quantum Karya Bersama.

Menghasilkan tiga berkas di `templates/`:
  1. Laporan-Pemasukan-Pengeluaran.xlsx
  2. Form-Kas.xlsx
  3. Buku-Kas-Sederhana.xlsx

Semua berkas memakai kop yang sama (lambang + identitas perusahaan) dan sudah
berisi rumus penjumlahan/saldo berjalan, jadi tinggal diisi barisnya.

Jalankan: python3 scripts/generate-office-templates.py
"""
from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
LOGO = os.path.join(ROOT, "assets", "logo-quantum.png")
OUT_DIR = os.path.join(ROOT, "templates")

# Warna identitas dari papan nama bengkel.
GOLD = "F2B705"
BLUE = "1B4FD8"
RED = "E0202B"
DARK = "0F172A"
GREY = "F1F5F9"

COMPANY = "CV. QUANTUM KARYA BERSAMA"
BUSINESS = "KAROSERI · BODY REPAIR · SERVICE MOBIL"
ADDRESS = "Kp. Tenjo Laut No. 1 RT 01/01, Desa Sukakarya, Kec. Sukakarya, Kabupaten Bekasi"
PHONE = "WhatsApp 0858-8669-2214"

FONT = "Arial"
RUPIAH = '#,##0;(#,##0);"-"'

thin = Side(style="thin", color="94A3B8")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def draw_letterhead(ws, last_col: int, title: str, subtitle: str = "") -> int:
    """Gambar kop surat di baris 1-6. Mengembalikan baris pertama setelah kop."""
    last = get_column_letter(last_col)

    # Baris judul perusahaan di sebelah lambang.
    ws.merge_cells(f"C1:{last}1")
    ws["C1"] = COMPANY
    ws["C1"].font = Font(name=FONT, size=16, bold=True, color=DARK)
    ws["C1"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells(f"C2:{last}2")
    ws["C2"] = BUSINESS
    ws["C2"].font = Font(name=FONT, size=9, bold=True, color=BLUE)
    ws["C2"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells(f"C3:{last}3")
    ws["C3"] = ADDRESS
    ws["C3"].font = Font(name=FONT, size=8, color="475569")
    ws["C3"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells(f"C4:{last}4")
    ws["C4"] = PHONE
    ws["C4"].font = Font(name=FONT, size=8, color="475569")
    ws["C4"].alignment = Alignment(horizontal="left", vertical="center")

    for row in (1, 2, 3, 4):
        ws.row_dimensions[row].height = 18

    # Garis emas pemisah kop, meniru aksen pada papan nama.
    ws.merge_cells(f"A5:{last}5")
    ws["A5"].fill = PatternFill("solid", fgColor=GOLD)
    ws.row_dimensions[5].height = 6

    # Bilah judul dokumen.
    ws.merge_cells(f"A6:{last}6")
    ws["A6"] = title
    ws["A6"].font = Font(name=FONT, size=13, bold=True, color="FFFFFF")
    ws["A6"].fill = PatternFill("solid", fgColor=DARK)
    ws["A6"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[6].height = 26

    next_row = 7
    if subtitle:
        ws.merge_cells(f"A7:{last}7")
        ws["A7"] = subtitle
        ws["A7"].font = Font(name=FONT, size=9, italic=True, color="475569")
        ws["A7"].alignment = Alignment(horizontal="center")
        next_row = 8

    if os.path.exists(LOGO):
        img = XLImage(LOGO)
        img.height = 62
        img.width = 62
        ws.add_image(img, "A1")

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 5
    return next_row + 1


def header_row(ws, row: int, labels: list[str], widths: list[int]) -> None:
    for idx, (label, width) in enumerate(zip(labels, widths), start=1):
        cell = ws.cell(row=row, column=idx, value=label)
        cell.font = Font(name=FONT, size=9, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.row_dimensions[row].height = 28


def body_rows(ws, first: int, count: int, cols: int, money_cols: list[int]) -> None:
    for row in range(first, first + count):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = BORDER
            cell.font = Font(name=FONT, size=10)
            if col in money_cols:
                cell.number_format = RUPIAH
                cell.alignment = Alignment(horizontal="right")
        ws.row_dimensions[row].height = 18


def legend(ws, row: int, last_col: int, text: str) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(name=FONT, size=8, italic=True, color="64748B")
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 24


def build_laporan() -> None:
    """Laporan pemasukan & pengeluaran — dua blok berdampingan seperti form cetak."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan"
    ws.sheet_view.showGridLines = False

    start = draw_letterhead(ws, 7, "LAPORAN PEMASUKAN & PENGELUARAN")

    ws.cell(row=start, column=1, value="Periode:").font = Font(name=FONT, size=10, bold=True)
    ws.cell(row=start, column=2, value="").border = BORDER
    ws.cell(row=start, column=3, value="s/d").font = Font(name=FONT, size=10, bold=True)
    ws.cell(row=start, column=4, value="").border = BORDER
    start += 2

    # Judul dua blok.
    ws.merge_cells(start_row=start, start_column=1, end_row=start, end_column=3)
    ws.cell(row=start, column=1, value="A. PEMASUKAN").font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
    ws.cell(row=start, column=1).fill = PatternFill("solid", fgColor=BLUE)
    ws.cell(row=start, column=1).alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=start, start_column=5, end_row=start, end_column=7)
    ws.cell(row=start, column=5, value="B. PENGELUARAN").font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
    ws.cell(row=start, column=5).fill = PatternFill("solid", fgColor=RED)
    ws.cell(row=start, column=5).alignment = Alignment(horizontal="center")
    ws.row_dimensions[start].height = 22

    head = start + 1
    for col, label, width in (
        (1, "TANGGAL", 13), (2, "KETERANGAN", 30), (3, "JUMLAH (Rp)", 16),
        (4, "", 3),
        (5, "TANGGAL", 13), (6, "KETERANGAN", 30), (7, "JUMLAH (Rp)", 16),
    ):
        if not label:
            ws.column_dimensions[get_column_letter(col)].width = width
            continue
        cell = ws.cell(row=head, column=col, value=label)
        cell.font = Font(name=FONT, size=9, bold=True)
        cell.fill = PatternFill("solid", fgColor=GREY)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col)].width = width

    rows = 20
    first = head + 1
    last = first + rows - 1
    for row in range(first, last + 1):
        for col in (1, 2, 3, 5, 6, 7):
            cell = ws.cell(row=row, column=col)
            cell.border = BORDER
            cell.font = Font(name=FONT, size=10)
            if col in (3, 7):
                cell.number_format = RUPIAH
                cell.alignment = Alignment(horizontal="right")

    total = last + 1
    ws.merge_cells(start_row=total, start_column=1, end_row=total, end_column=2)
    ws.cell(row=total, column=1, value="TOTAL PEMASUKAN").font = Font(name=FONT, size=10, bold=True)
    ws.cell(row=total, column=1).alignment = Alignment(horizontal="right")
    ws.cell(row=total, column=3, value=f"=SUM(C{first}:C{last})")

    ws.merge_cells(start_row=total, start_column=5, end_row=total, end_column=6)
    ws.cell(row=total, column=5, value="TOTAL PENGELUARAN").font = Font(name=FONT, size=10, bold=True)
    ws.cell(row=total, column=5).alignment = Alignment(horizontal="right")
    ws.cell(row=total, column=7, value=f"=SUM(G{first}:G{last})")

    for col in (3, 7):
        cell = ws.cell(row=total, column=col)
        cell.font = Font(name=FONT, size=11, bold=True)
        cell.number_format = RUPIAH
        cell.fill = PatternFill("solid", fgColor=GREY)
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="right")

    saldo = total + 2
    ws.merge_cells(start_row=saldo, start_column=1, end_row=saldo, end_column=2)
    ws.cell(row=saldo, column=1, value="SALDO BERSIH").font = Font(name=FONT, size=12, bold=True, color="FFFFFF")
    ws.cell(row=saldo, column=1).fill = PatternFill("solid", fgColor=GOLD)
    ws.cell(row=saldo, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=saldo, column=3, value=f"=C{total}-G{total}")
    ws.cell(row=saldo, column=3).font = Font(name=FONT, size=12, bold=True)
    ws.cell(row=saldo, column=3).number_format = RUPIAH
    ws.cell(row=saldo, column=3).border = BORDER
    ws.cell(row=saldo, column=3).alignment = Alignment(horizontal="right")
    ws.cell(row=saldo, column=4, value="( Pemasukan − Pengeluaran )").font = Font(name=FONT, size=8, italic=True)
    ws.row_dimensions[saldo].height = 24

    legend(ws, saldo + 2, 7,
           "Cara pakai: isi tanggal, keterangan, dan jumlah pada baris kosong. Baris TOTAL dan SALDO BERSIH "
           "terisi otomatis oleh rumus — jangan diketik manual. Butuh baris tambahan? Sisipkan baris di "
           "tengah tabel supaya rumus penjumlahannya ikut melebar.")

    ws.page_setup.orientation = "landscape"
    ws.print_area = f"A1:G{saldo + 2}"
    wb.save(os.path.join(OUT_DIR, "Laporan-Pemasukan-Pengeluaran.xlsx"))


def build_form_kas() -> None:
    """Form kas: penerimaan, pengeluaran, saldo berjalan, plus blok tanda tangan."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Form Kas"
    ws.sheet_view.showGridLines = False

    start = draw_letterhead(ws, 5, "FORM KAS")

    for offset, label in enumerate(("Tanggal", "Bulan", "Tahun")):
        row = start + offset
        ws.cell(row=row, column=4, value=f"{label}:").font = Font(name=FONT, size=9, bold=True)
        ws.cell(row=row, column=5).border = BORDER
    start += 4

    header_row(ws, start, ["NO.", "URAIAN", "PENERIMAAN (Rp)", "PENGELUARAN (Rp)", "SALDO (Rp)"],
               [6, 40, 18, 18, 18])

    rows = 15
    first = start + 1
    last = first + rows - 1
    body_rows(ws, first, rows, 5, [3, 4, 5])

    for idx, row in enumerate(range(first, last + 1), start=1):
        ws.cell(row=row, column=1, value=idx).alignment = Alignment(horizontal="center")
        # Saldo berjalan: baris pertama berdiri sendiri, berikutnya menumpuk ke atasnya.
        if row == first:
            ws.cell(row=row, column=5, value=f"=C{row}-D{row}")
        else:
            ws.cell(row=row, column=5, value=f"=E{row - 1}+C{row}-D{row}")

    total = last + 1
    ws.merge_cells(start_row=total, start_column=1, end_row=total, end_column=2)
    ws.cell(row=total, column=1, value="TOTAL").font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
    ws.cell(row=total, column=1).fill = PatternFill("solid", fgColor=DARK)
    ws.cell(row=total, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=total, column=3, value=f"=SUM(C{first}:C{last})")
    ws.cell(row=total, column=4, value=f"=SUM(D{first}:D{last})")
    ws.cell(row=total, column=5, value=f"=E{last}")
    for col in (3, 4, 5):
        cell = ws.cell(row=total, column=col)
        cell.font = Font(name=FONT, size=11, bold=True)
        cell.number_format = RUPIAH
        cell.border = BORDER
        cell.fill = PatternFill("solid", fgColor=GREY)
        cell.alignment = Alignment(horizontal="right")
    ws.row_dimensions[total].height = 22

    sign = total + 2
    for col, label in ((1, "Dibuat oleh,"), (3, "Diperiksa oleh,"), (5, "Disetujui oleh,")):
        ws.cell(row=sign, column=col, value=label).font = Font(name=FONT, size=9)
        ws.cell(row=sign, column=col).alignment = Alignment(horizontal="center")
        ws.cell(row=sign + 4, column=col, value="(________________)").font = Font(name=FONT, size=9)
        ws.cell(row=sign + 4, column=col).alignment = Alignment(horizontal="center")

    legend(ws, sign + 6, 5,
           "Cara pakai: isi kolom URAIAN, PENERIMAAN, dan PENGELUARAN. Kolom SALDO dan baris TOTAL "
           "dihitung otomatis oleh rumus — jangan diketik manual.")

    ws.print_area = f"A1:E{sign + 6}"
    wb.save(os.path.join(OUT_DIR, "Form-Kas.xlsx"))


def build_buku_kas() -> None:
    """Buku kas sederhana: satu baris per transaksi dengan saldo berjalan."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Buku Kas"
    ws.sheet_view.showGridLines = False

    start = draw_letterhead(ws, 5, "BUKU KAS SEDERHANA", "Bulan: ______________  Tahun: __________")

    header_row(ws, start, ["TANGGAL", "KETERANGAN", "PENERIMAAN (Rp)", "PENGELUARAN (Rp)", "SALDO (Rp)"],
               [14, 40, 18, 18, 18])

    rows = 30
    first = start + 1
    last = first + rows - 1
    body_rows(ws, first, rows, 5, [3, 4, 5])

    for row in range(first, last + 1):
        if row == first:
            ws.cell(row=row, column=5, value=f"=C{row}-D{row}")
        else:
            ws.cell(row=row, column=5, value=f"=E{row - 1}+C{row}-D{row}")

    total = last + 1
    ws.merge_cells(start_row=total, start_column=1, end_row=total, end_column=2)
    ws.cell(row=total, column=1, value="TOTAL").font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
    ws.cell(row=total, column=1).fill = PatternFill("solid", fgColor=DARK)
    ws.cell(row=total, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=total, column=3, value=f"=SUM(C{first}:C{last})")
    ws.cell(row=total, column=4, value=f"=SUM(D{first}:D{last})")
    ws.cell(row=total, column=5, value=f"=E{last}")
    for col in (3, 4, 5):
        cell = ws.cell(row=total, column=col)
        cell.font = Font(name=FONT, size=11, bold=True)
        cell.number_format = RUPIAH
        cell.border = BORDER
        cell.fill = PatternFill("solid", fgColor=GREY)
        cell.alignment = Alignment(horizontal="right")
    ws.row_dimensions[total].height = 22

    legend(ws, total + 2, 5,
           "Cara pakai: satu baris untuk satu transaksi, urut tanggal. Kolom SALDO menumpuk otomatis dari "
           "baris di atasnya; baris TOTAL menjumlah sendiri. Isi hanya kolom TANGGAL, KETERANGAN, "
           "PENERIMAAN, dan PENGELUARAN.")

    ws.print_area = f"A1:E{total + 2}"
    wb.save(os.path.join(OUT_DIR, "Buku-Kas-Sederhana.xlsx"))


def main() -> None:
    os.makedirs(OUT_DIR, exist_ok=True)
    build_laporan()
    build_form_kas()
    build_buku_kas()
    print("Template Excel dibuat di templates/:")
    for name in sorted(os.listdir(OUT_DIR)):
        print(f"  - {name}")


if __name__ == "__main__":
    main()
